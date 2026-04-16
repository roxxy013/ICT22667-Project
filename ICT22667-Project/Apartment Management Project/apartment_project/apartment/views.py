from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.db import models as django_models
from django.db.models import Sum, Count, Q, Subquery, OuterRef, Value
from django.db.models.functions import Concat
from django.contrib.auth.decorators import login_required
from django.core.mail import send_mail
from django.template.loader import render_to_string
from decimal import Decimal
from .models import Tenant, Room, Contract, Invoice, MonthlyBill, Utility, Fine, Maintenance, Booking, EmployeeSalary
from .forms  import TenantForm, RoomForm, ContractForm, InvoiceForm, UtilityForm, PaymentForm, FineForm, MaintenanceForm, BookingForm, EmployeeSalaryForm
from .decorators import role_required
import datetime
from django.core.mail import send_mass_mail
import time

def get_user_building(user):
    from .middleware import get_user_role
    role = get_user_role(user)
    if role in ['MANAGER', 'METER'] and hasattr(user, 'userprofile') and user.userprofile.Building_No:
        return user.userprofile.Building_No
    return None

# ==================== DASHBOARD ====================

@login_required
@role_required('ADMIN', 'MANAGER', 'METER')
def dashboard(request):
    from .middleware import get_user_role
    if get_user_role(request.user) == 'METER':
        return redirect('meter_input')
    import datetime
    today = datetime.date.today()
    building = get_user_building(request.user)
    rooms = Room.objects.all()
    if building: rooms = rooms.filter(Building_No=building)
    rooms = rooms.order_by('Building_No', 'Floor', 'Room_Number')

    Invoice.objects.filter(
        Status='รอชำระ',
        Due_Date__lt=today
    ).update(Status='เกินกำหนด')

    # ห้องที่มี invoice เกินกำหนด (เฉพาะสัญญาที่ยัง active) → สีแดง
    overdue_qs = Invoice.objects.filter(Status='เกินกำหนด', Contract_ID__Status='ใช้งาน')
    if building: overdue_qs = overdue_qs.filter(Contract_ID__Room_ID__Building_No=building)
    overdue_room_ids = list(overdue_qs.values_list('Contract_ID__Room_ID', flat=True))

    # ห้องที่มี invoice รอชำระ (ยังไม่เกิน, เฉพาะสัญญา active) → แสดง $
    unpaid_qs = Invoice.objects.filter(Status='รอชำระ', Contract_ID__Status='ใช้งาน')
    if building: unpaid_qs = unpaid_qs.filter(Contract_ID__Room_ID__Building_No=building)
    unpaid_room_ids = list(unpaid_qs.values_list('Contract_ID__Room_ID', flat=True))

    # ห้องที่มีแจ้งซ่อมค้าง → แสดง 🔧
    repair_qs = Maintenance.objects.exclude(Status='ซ่อมเสร็จ')
    if building: repair_qs = repair_qs.filter(Room_ID__Building_No=building)
    repair_room_ids = list(repair_qs.values_list('Room_ID', flat=True))


    # --- นับตาม "สีจริง" ที่แสดงใน badge ---
    count_white    = rooms.filter(Status='ว่าง', Status_Flag='ปกติ').count()
    count_pin      = rooms.filter(Status='ว่าง', Status_Flag='จอง').count()
    count_broom    = rooms.filter(Status_Flag='รอทำความสะอาด').count()
    
    # มีผู้เช่า: นับทุกคนที่ Status เป็น 'มีผู้เช่า' (รวมปกติ, แจ้งย้ายออก, และเกินกำหนด)
    count_blue     = rooms.filter(Status='มีผู้เช่า').count()
    
    # แจ้งย้ายออก: นับเฉพาะคนที่มีแผนจะย้ายออก (แต่ยังไม่ลด count_blue)
    count_yellow   = rooms.filter(Status='มีผู้เช่า', Status_Flag='แจ้งย้ายออก').count()
    
    # เกินกำหนด: นับตาม Invoice (แต่ยังไม่ลด count_blue)
    count_red      = rooms.filter(Room_ID__in=overdue_room_ids).count()
    
    count_black    = rooms.filter(Status='ซ่อมบำรุง').count()
    count_repair   = len(set(repair_room_ids))
    count_unpaid   = len(set(unpaid_room_ids))

    # ห้องที่รอทำความสะอาด (ใช้แสดง icon บน badge)
    clean_room_ids = list(rooms.filter(
        Status_Flag='รอทำความสะอาด'
    ).values_list('Room_ID', flat=True))

    context = {
        'rooms':            rooms,
        'overdue_room_ids': overdue_room_ids,
        'unpaid_room_ids':  unpaid_room_ids,
        'repair_room_ids':  repair_room_ids,
        'clean_room_ids':   clean_room_ids,
        # summary cards
        'total_rooms':   rooms.count(),
        'count_white':   count_white,
        'count_pin':     count_pin,
        'count_broom':   count_broom,
        'count_blue':    count_blue,
        'count_yellow':  count_yellow,
        'count_red':     count_red,
        'count_black':   count_black,
        'count_repair':  count_repair,
        'count_unpaid':  count_unpaid,
    }
    return render(request, 'apartment/dashboard.html', context)


# ==================== TENANT ====================

@login_required
@role_required('ADMIN', 'MANAGER')
def tenant_list(request):
    query    = request.GET.get('q', '')
    building = request.GET.get('building', '')
    floor    = request.GET.get('floor', '')

    # ดึงค่าห้องพักปัจจุบันของแต่ละ Tenant จาก Contract ล่าสุดที่ยัง 'ใช้งาน'
    active_contract = Contract.objects.filter(
        Tenant_ID=OuterRef('pk'),
        Status='ใช้งาน'
    )
    
    tenants = Tenant.objects.annotate(
        room_number=Subquery(active_contract.values('Room_ID__Room_Number')[:1]),
        building_no=Subquery(active_contract.values('Room_ID__Building_No')[:1]),
        floor_no=Subquery(active_contract.values('Room_ID__Floor')[:1])
    )

    # Manager restriction
    user_building = get_user_building(request.user)
    if user_building:
        building = user_building # Force lock
        tenants = tenants.filter(building_no=user_building)

    # กรองตามการค้นหา (รวมชื่อ นามสกุล และเลขห้อง)
    if query:
        tenants = tenants.annotate(
            full_name=Concat('First_Name', Value(' '), 'Last_Name')
        ).filter(
            Q(First_Name__icontains=query) | 
            Q(Last_Name__icontains=query) |
            Q(full_name__icontains=query) |
            Q(room_number__icontains=query)
        )
    
    if building:
        tenants = tenants.filter(building_no=building)
    if floor:
        tenants = tenants.filter(floor_no=floor)

    # นำค่าตึกและชั้นทั้งหมดเพื่อไปใส่ใน Dropdown ตัวกรอง (ถ้า manager ล็อกอิน จะแสดงแค่ตึกตัวเอง)
    buildings_qs = Room.objects.values_list('Building_No', flat=True).distinct().order_by('Building_No')
    if user_building:
        buildings_qs = [user_building]
    
    floors = Room.objects.values_list('Floor', flat=True).distinct().order_by('Floor')

    # เรียงลำดับตามตึก ชั้น และเลขห้อง
    tenants = tenants.order_by('building_no', 'floor_no', 'room_number', 'First_Name')

    return render(request, 'apartment/tenant/list.html', {
        'tenants':   tenants, 
        'query':     query,
        'building':  building,
        'floor':     floor,
        'buildings': buildings_qs,
        'floors':    floors,
        'user_building': user_building,
    })

@login_required
@role_required('ADMIN', 'MANAGER')
def tenant_create(request):
    form = TenantForm(request.POST or None)
    if form.is_valid():
        form.save()
        return redirect('tenant_list')
    return render(request, 'apartment/tenant/form.html', {'form': form, 'title': 'เพิ่มผู้เช่า'})

@login_required
@role_required('ADMIN', 'MANAGER')
def tenant_edit(request, pk):
    tenant = get_object_or_404(Tenant, pk=pk)
    form   = TenantForm(request.POST or None, instance=tenant)
    if form.is_valid():
        form.save()
        return redirect('tenant_list')
    return render(request, 'apartment/tenant/form.html', {'form': form, 'title': 'แก้ไขผู้เช่า'})

@login_required
@role_required('ADMIN')
def tenant_delete(request, pk):
    tenant = get_object_or_404(Tenant, pk=pk)
    if request.method == 'POST':
        tenant.delete()
        return redirect('tenant_list')
    return render(request, 'apartment/tenant/confirm_delete.html', {'object': tenant, 'title': 'ลบผู้เช่า'})


# ==================== ROOM ====================

@login_required
@role_required('ADMIN', 'MANAGER')
def room_list(request):
    building = request.GET.get('building', '')
    floor    = request.GET.get('floor', '')
    
    # Manager restriction
    user_building = get_user_building(request.user)
    if user_building:
        building = user_building

    rooms = Room.objects.all()
    if building: rooms = rooms.filter(Building_No=building)
    if floor:    rooms = rooms.filter(Floor=floor)
    
    rooms = rooms.order_by('Building_No', 'Floor', 'Room_Number')

    # รายการอาคารและชั้นสำหรับ Filter
    buildings_qs = Room.objects.values_list('Building_No', flat=True).distinct().order_by('Building_No')
    if user_building:
        buildings_qs = [user_building]
    
    floors = Room.objects.values_list('Floor', flat=True).distinct().order_by('Floor')

    return render(request, 'apartment/room/list.html', {
        'rooms': rooms,
        'building': building,
        'floor': floor,
        'buildings': buildings_qs,
        'floors': floors,
        'user_building': user_building,
    })

@login_required
@role_required('ADMIN')
def room_create(request):
    form = RoomForm(request.POST or None)
    if form.is_valid():
        form.save()
        return redirect('room_list')
    return render(request, 'apartment/room/form.html', {'form': form, 'title': 'เพิ่มห้องพัก'})

@login_required
@role_required('ADMIN')
def room_edit(request, pk):
    room = get_object_or_404(Room, pk=pk)
    form = RoomForm(request.POST or None, instance=room)
    if form.is_valid():
        form.save()
        return redirect('room_list')
    return render(request, 'apartment/room/form.html', {'form': form, 'title': 'แก้ไขห้องพัก'})

@login_required
@role_required('ADMIN')
def room_delete(request, pk):
    room = get_object_or_404(Room, pk=pk)
    if request.method == 'POST':
        room.delete()
        return redirect('room_list')
    return render(request, 'apartment/room/confirm_delete.html', {'object': room, 'title': 'ลบห้องพัก'})

@login_required
@role_required('ADMIN', 'MANAGER')
def room_detail(request, pk):
    room = get_object_or_404(Room, pk=pk)

    # 1. สัญญาปัจจุบัน (ถ้ามี)
    active_contract = Contract.objects.filter(
        Room_ID=room, Status='ใช้งาน'
    ).select_related('Tenant_ID').first()

    # 2. ประวัติสัญญาทั้งหมดของห้องนี้ (เรียงจากใหม่ไปเก่า)
    all_contracts = Contract.objects.filter(
        Room_ID=room
    ).select_related('Tenant_ID').order_by('-Contract_ID')

    # 3. ดึงแจ้งซ่อมของห้องนี้
    maintenances = Maintenance.objects.filter(
        Room_ID=room
    ).order_by('-Report_Date')

    # 4. ดึงการจองที่รอยืนยัน
    booking = Booking.objects.filter(
        Room_ID=room, Status='รอยืนยัน'
    ).first()

    return render(request, 'apartment/room/detail.html', {
        'room':            room,
        'active_contract': active_contract,
        'all_contracts':   all_contracts, # ส่งสัญญาทั้งหมดไปแยก Section ใน Template
        'maintenances':    maintenances,
        'booking':         booking,
    })
# ==================== CONTRACT ====================

@login_required
@role_required('ADMIN', 'MANAGER')
def contract_list(request):
    building = get_user_building(request.user)
    contracts = Contract.objects.select_related('Tenant_ID', 'Room_ID').all()
    
    # --- Filter ---
    q      = request.GET.get('q', '')
    status = request.GET.get('status', '')

    if building:
        contracts = contracts.filter(Room_ID__Building_No=building)
    
    if q:
        contracts = contracts.annotate(
            full_name=Concat('Tenant_ID__First_Name', Value(' '), 'Tenant_ID__Last_Name')
        ).filter(
            Q(Tenant_ID__First_Name__icontains=q) |
            Q(Tenant_ID__Last_Name__icontains=q)  |
            Q(full_name__icontains=q) |
            Q(Room_ID__Room_Number__icontains=q)
        )
    
    if status:
        contracts = contracts.filter(Status=status)

    contracts = contracts.order_by('-Contract_ID') # เรียงใหม่ล่าสุดขึ้นก่อน

    # รายการสถานะสำหรับ Filter
    status_choices = Contract.objects.values_list('Status', flat=True).distinct()

    return render(request, 'apartment/contract/list.html', {
        'contracts': contracts,
        'q': q,
        'status': status,
        'status_choices': status_choices,
    })

@login_required
@role_required('ADMIN', 'MANAGER')
def contract_create(request, room_pk=None):
    # ค่าเริ่มต้นสำหรับสัญญา
    initial_data = {
        'Rent_Price':       4000,
        'Deposit':          4000,
        'Deposit_Advance':  2000,
        'Water_Cost_Unit':  18,
        'Elec_Cost_Unit':   8,
        'Status':           'ใช้งาน',
    }

    if room_pk:
        room = get_object_or_404(Room, pk=room_pk, Status='ว่าง')
        initial_data['Room_ID'] = room
        # ดึงหน่วยมิเตอร์ล่าสุดของห้องนี้ (จาก Utility เก่า หรือ Contract เก่า)
        latest_u = Utility.objects.filter(Room_ID=room).order_by('-Bill_Month').first()
        if latest_u:
            initial_data['Water_Meter_Start'] = latest_u.Water_Unit_After
            initial_data['Elec_Meter_Start']  = latest_u.Elec_Unit_After
        else:
            latest_c = Contract.objects.filter(Room_ID=room).order_by('-Contract_ID').first()
            if latest_c:
                initial_data['Water_Meter_Start'] = latest_c.Water_Meter_Start
                initial_data['Elec_Meter_Start']  = latest_c.Elec_Meter_Start

    form = ContractForm(request.POST or None, initial=initial_data)

    building = get_user_building(request.user)
    if room_pk:
        qs = Room.objects.filter(pk=room_pk)
    else:
        qs = Room.objects.filter(Status='ว่าง')
        
    if building: 
        qs = qs.filter(Building_No=building)
        
    form.fields['Room_ID'].queryset = qs

    if form.is_valid():
        contract = form.save(commit=False)
        contract.save()
        # อัปเดตสถานะห้องเป็น "มีผู้เช่า"
        room        = contract.Room_ID
        room.Status = 'มีผู้เช่า'
        room.save()
        return redirect('contract_print', pk=contract.Contract_ID)  # ไปหน้าพิมพ์สัญญาเลย

    return render(request, 'apartment/contract/form.html', {
        'form':  form,
        'title': 'สร้างสัญญาเข้าพัก',
    })


@login_required
@role_required('ADMIN', 'MANAGER')
def contract_edit(request, pk):
    contract = get_object_or_404(Contract, pk=pk)
    form     = ContractForm(request.POST or None, instance=contract)
    if form.is_valid():
        form.save()
        return redirect('contract_list')
    return render(request, 'apartment/contract/form.html', {
        'form':  form,
        'title': 'แก้ไขสัญญาเข้าพัก',
    })


@login_required
@role_required('ADMIN', 'MANAGER')
def contract_print(request, pk):
    # หน้าพิมพ์สัญญา
    contract = get_object_or_404(Contract, pk=pk)
    return render(request, 'apartment/contract/print.html', {'contract': contract})

@login_required
@role_required('ADMIN')
def contract_delete(request, pk):
    contract = get_object_or_404(Contract, pk=pk)
    if request.method == 'POST':
        contract.delete()
        return redirect('contract_list')
    return render(request, 'apartment/contract/confirm_delete.html', {'object': contract, 'title': 'ลบสัญญาเช่า'})
# ==================== INVOICE ====================

@login_required
@role_required('ADMIN', 'MANAGER')
def invoice_list(request):
    import datetime
    today = datetime.date.today()

    # อัปเดตสถานะเกินกำหนดอัตโนมัติ
    Invoice.objects.filter(
        Status='รอชำระ', Due_Date__lt=today
    ).update(Status='เกินกำหนด')

    invoices = Invoice.objects.select_related(
        'Contract_ID__Tenant_ID', 'Contract_ID__Room_ID'
    ).all()

    # --- Filter ---
    q        = request.GET.get('q', '')
    month    = request.GET.get('month', '')
    year     = request.GET.get('year', str(today.year))
    status   = request.GET.get('status', '')
    building = request.GET.get('building', '')
    sort     = request.GET.get('sort', 'room')

    # Manager restriction: ล็อคอาคารถ้าเป็น Manager
    user_building = get_user_building(request.user)
    if user_building:
        building = user_building  # Force lock to their building

    if q:
        invoices = invoices.annotate(
            full_name=Concat('Contract_ID__Tenant_ID__First_Name', Value(' '), 'Contract_ID__Tenant_ID__Last_Name')
        ).filter(
            Q(Contract_ID__Tenant_ID__First_Name__icontains=q) |
            Q(Contract_ID__Tenant_ID__Last_Name__icontains=q)  |
            Q(full_name__icontains=q) |
            Q(Contract_ID__Room_ID__Room_Number__icontains=q)
        )
    if month:
        invoices = invoices.filter(Billing_Date__month=month)
    if year:
        invoices = invoices.filter(Billing_Date__year=year)
    if status:
        invoices = invoices.filter(Status=status)
    if building:
        invoices = invoices.filter(Contract_ID__Room_ID__Building_No=building)

    # --- Sort ---
    if sort == 'amount_desc':
        invoices = invoices.order_by('-Grand_Total')
    elif sort == 'amount_asc':
        invoices = invoices.order_by('Grand_Total')
    elif sort == 'paid_date':
        invoices = invoices.order_by('-Paid_Date')
    else:
        invoices = invoices.order_by('Contract_ID__Room_ID__Room_Number')

    # dropdown เดือน/ปี
    months_th = [
        (1,'มกราคม'),(2,'กุมภาพันธ์'),(3,'มีนาคม'),(4,'เมษายน'),
        (5,'พฤษภาคม'),(6,'มิถุนายน'),(7,'กรกฎาคม'),(8,'สิงหาคม'),
        (9,'กันยายน'),(10,'ตุลาคม'),(11,'พฤศจิกายน'),(12,'ธันวาคม'),
    ]
    
    # รายการอาคารสำหรับ Filter
    buildings_qs = Room.objects.values_list('Building_No', flat=True).distinct().order_by('Building_No')
    if user_building:
        buildings_qs = [user_building]

    all_years = Invoice.objects.dates('Billing_Date', 'year', order='DESC')
    # กรองเอาเฉพาะปีที่มากกว่า 0 อย่างเข้มงวด
    year_list = sorted([y.year for y in all_years if y and y.year > 0], reverse=True)
    if not year_list:
        year_list = [today.year]

    return render(request, 'apartment/invoice/list.html', {
        'invoices':   invoices,
        'q':          q,
        'month':      int(month) if month else '',
        'year':       int(year) if (year and year != '0') else '',
        'status':     status,
        'building':   building,
        'sort':       sort,
        'months_th':  months_th,
        'years':      year_list,
        'buildings':  buildings_qs,
        'user_building': user_building, # เพื่อเช็คใน template
    })


@login_required
@role_required('ADMIN', 'MANAGER')
def invoice_create(request):
    contract_id = request.GET.get('contract_id')
    initial_invoice = {}
    initial_utility = {}

    today = datetime.date.today()
    bill_month = today.replace(day=1)

    if contract_id:
        try:
            contract = Contract.objects.get(pk=contract_id)
            initial_invoice['Contract_ID'] = contract.Contract_ID
            initial_utility['Water_Cost_Unit'] = contract.Water_Cost_Unit
            initial_utility['Elec_Cost_Unit'] = contract.Elec_Cost_Unit
            initial_utility['Bill_Month'] = bill_month.strftime('%Y-%m-%d')

            # ดึงข้อมูล Utility ล่าสุดของสัญญาเพื่อดึงหน่วยที่ใช้ (ถ้ามี)
            utility = Utility.objects.filter(Room_ID=contract.Room_ID).order_by('-Bill_Month').first()
            if utility:
                initial_utility['Water_Unit_Used'] = utility.Water_Unit_Used
                initial_utility['Elec_Unit_Used'] = utility.Elec_Unit_Used
                # ถ้า Utility ล่าสุดเป็นของเดือนนี้ ให้ใช้เดือนตามนั้นเลย
                initial_utility['Bill_Month'] = utility.Bill_Month.strftime('%Y-%m-%d')
            else:
                initial_utility['Water_Unit_Used'] = 0
                initial_utility['Elec_Unit_Used'] = 0
        except Contract.DoesNotExist:
            pass

    invoice_form = InvoiceForm(request.POST or None, initial=initial_invoice)
    utility_form = UtilityForm(request.POST or None, initial=initial_utility)

    # คำนวณ Due_Date อัตโนมัติ = วันที่ 5 ของเดือนถัดไป

    today = datetime.date.today() #ดึงวันนี้เข้า
    billing_date  = today.replace(day=25)  # วันที่ 25 ของเดือนนี้
    next_month    = (today.replace(day=1) + datetime.timedelta(days=32)).replace(day=1)
    due_date      = next_month.replace(day=5)

    if request.method == 'POST':
        if invoice_form.is_valid() and utility_form.is_valid():
            # 1. ดึงข้อมูลจาก contract
            contract = invoice_form.cleaned_data['Contract_ID']

            # 2. สร้าง Invoice ก่อน (ยังไม่คำนวณ Grand_Total)
            invoice = invoice_form.save(commit=False)
            invoice.Grand_Total = 0
            invoice.save()

            # 3. สร้าง MonthlyBill (ค่าเช่าจาก contract)
            monthly_bill = MonthlyBill.objects.create(
                Invoice_ID = invoice,
                Bill_Month = invoice_form.cleaned_data['Billing_Date'],
                Amount     = contract.Rent_Price,
            )

            # 4. สร้าง Utility (คำนวณอัตโนมัติ)
            u = utility_form.save(commit=False)
            u.Invoice_ID      = invoice
            u.Room_ID         = contract.Room_ID
            u.Water_Unit_Before = 0
            u.Water_Unit_After = u.Water_Unit_Used
            u.Water_Total     = u.Water_Unit_Used * Decimal(u.Water_Cost_Unit)
            u.Elec_Total      = Decimal(u.Elec_Unit_Used) * Decimal(u.Elec_Cost_Unit)
            u.save()

            # 5. คำนวณ Grand_Total แล้ว save กลับ
            invoice.Grand_Total = monthly_bill.Amount + u.Water_Total + u.Elec_Total
            invoice.save()

            return redirect('invoice_detail', pk=invoice.Invoice_ID)

    return render(request, 'apartment/invoice/form.html', {
        'invoice_form': invoice_form,
        'utility_form': utility_form,
        'title': 'ออกใบแจ้งหนี้',
    })


@login_required
@role_required('ADMIN', 'MANAGER')
def invoice_detail(request, pk):
    invoice      = get_object_or_404(Invoice, pk=pk)
    monthly_bill = MonthlyBill.objects.filter(Invoice_ID=invoice).first()
    utility      = Utility.objects.filter(Invoice_ID=invoice).first()
    fines        = Fine.objects.filter(Invoice_ID=invoice)
    fine_form    = FineForm(request.POST or None)

    # เพิ่มค่าปรับ
    if request.method == 'POST' and fine_form.is_valid():
        fine            = fine_form.save(commit=False)
        fine.Invoice_ID = invoice
        fine.save()
        # อัปเดต Grand_Total
        fine_total          = fines.aggregate(total=Sum('Amount'))['total'] or 0
        invoice.Grand_Total = (monthly_bill.Amount if monthly_bill else 0) + \
                              (utility.Water_Total + utility.Elec_Total if utility else 0) + \
                              fine_total + fine.Amount
        invoice.save()
        return redirect('invoice_detail', pk=pk)

    return render(request, 'apartment/invoice/detail.html', {
        'invoice':      invoice,
        'monthly_bill': monthly_bill,
        'utility':      utility,
        'fines':        fines,
        'fine_form':    fine_form,
    })


@login_required
@role_required('ADMIN', 'MANAGER')
def invoice_pay(request, pk):
    today   = datetime.date.today()
    invoice = get_object_or_404(Invoice, pk=pk)
    form    = PaymentForm(request.POST or None, instance=invoice)

    if form.is_valid():
        invoice           = form.save(commit=False)
        invoice.Paid_Date = today

        # ถ้าจ่ายหลัง due_date → จ่ายล่าช้า, ปกติ → ชำระแล้ว
        if invoice.Due_Date and today > invoice.Due_Date:
            invoice.Status = 'จ่ายล่าช้า'
        else:
            invoice.Status = 'ชำระแล้ว'

        invoice.save()
        return redirect('invoice_detail', pk=pk)

    return render(request, 'apartment/invoice/pay.html', {
        'form':    form,
        'invoice': invoice,
        'title':   'บันทึกชำระเงิน',
    })

@login_required
@role_required('ADMIN', 'MANAGER')
def invoice_extend(request, pk):
    today    = datetime.date.today()
    invoice  = get_object_or_404(Invoice, pk=pk)
    contract = invoice.Contract_ID
    tenant   = contract.Tenant_ID

    if request.method == 'POST':
        use_deposit = request.POST.get('use_deposit')

        if use_deposit == 'deposit':
            amount = contract.Deposit
            note   = 'โปะด้วยเงินประกันห้อง'
        else:
            amount = contract.Deposit_Advance
            note   = 'โปะด้วยเงินมัดจำ'

        Fine.objects.create(
            Invoice_ID = invoice,
            Reason     = note,
            Amount     = -amount,
            Fine_Date  = today,
        )

        fine_total   = Fine.objects.filter(
            Invoice_ID=invoice
        ).aggregate(t=Sum('Amount'))['t'] or 0

        monthly_bill = MonthlyBill.objects.filter(Invoice_ID=invoice).first()
        utility      = Utility.objects.filter(Invoice_ID=invoice).first()
        invoice.Grand_Total = (
            (monthly_bill.Amount if monthly_bill else 0) +
            (utility.Water_Total + utility.Elec_Total if utility else 0) +
            fine_total
        )

        # วันครบกำหนดใหม่ = วันที่ 5 ของเดือนถัดไปจากวันนี้
        next_m                    = (today + datetime.timedelta(days=10)).replace(day=1)
        invoice.Due_Date          = next_m.replace(day=5)
        invoice.Extended_Due_Date = invoice.Due_Date  # เก็บไว้ว่าเคยต่อเวลา
        invoice.Status            = 'ต่อเวลาชำระ'
        invoice.Paid_Date         = today              # บันทึกวันที่กดปุ่ม
        invoice.save()

        return redirect('invoice_detail', pk=pk)

    return render(request, 'apartment/invoice/extend.html', {
        'invoice':  invoice,
        'contract': contract,
        'tenant':   tenant,
    })


@login_required
@role_required('ADMIN', 'MANAGER')
def invoice_send_all_email(request):
    today = datetime.date.today()

    if request.method == 'POST':
        month = int(request.POST.get('month', today.month))
        year  = int(request.POST.get('year',  today.year))

        invoices = Invoice.objects.filter(
            Billing_Date__month = month,
            Billing_Date__year  = year,
        ).select_related(
            'Contract_ID__Tenant_ID',
            'Contract_ID__Room_ID'
        )

        sent    = 0
        failed  = 0
        no_mail = 0

        for i, invoice in enumerate(invoices):
            tenant = invoice.Contract_ID.Tenant_ID

            if not tenant.Email:
                no_mail += 1
                continue

            monthly_bill = MonthlyBill.objects.filter(Invoice_ID=invoice).first()
            utility      = Utility.objects.filter(Invoice_ID=invoice).first()
            fines        = Fine.objects.filter(Invoice_ID=invoice)

            try:
                email_body = render_to_string('apartment/invoice/email_body.html', {
                    'invoice':      invoice,
                    'monthly_bill': monthly_bill,
                    'utility':      utility,
                    'fines':        fines,
                    'tenant':       tenant,
                })
                send_mail(
                    subject        = f'ใบแจ้งหนี้ห้อง {invoice.Contract_ID.Room_ID} — {invoice.Billing_Date.strftime("%B %Y")}',
                    message        = '',
                    from_email     = None,
                    recipient_list = [tenant.Email],
                    html_message   = email_body,
                    fail_silently  = False,
                )
                sent += 1

                # พัก 0.5 วิทุก 10 ฉบับ ป้องกัน Gmail rate limit
                if sent % 10 == 0:
                    import time
                    time.sleep(0.5)

            except Exception:
                failed += 1

        return render(request, 'apartment/invoice/send_all_result.html', {
            'sent':    sent,
            'failed':  failed,
            'no_mail': no_mail,
            'month':   month,
            'year':    year,
        })

    # GET: หน้าเลือกเดือน
    months_th = [
        (1,'มกราคม'),(2,'กุมภาพันธ์'),(3,'มีนาคม'),(4,'เมษายน'),
        (5,'พฤษภาคม'),(6,'มิถุนายน'),(7,'กรกฎาคม'),(8,'สิงหาคม'),
        (9,'กันยายน'),(10,'ตุลาคม'),(11,'พฤศจิกายน'),(12,'ธันวาคม'),
    ]
    return render(request, 'apartment/invoice/send_all_confirm.html', {
        'months_th': months_th,
        'month':     today.month,
        'year':      today.year,
    })

def auto_generate_invoices():
    import datetime
    today = datetime.date.today()

    bill_month = today.replace(day=1)
    bill_date  = today.replace(day=25)
    next_m     = (today + datetime.timedelta(days=10)).replace(day=1)
    due_date   = next_m.replace(day=5)

    contracts = Contract.objects.filter(
        Status='ใช้งาน'
    ).select_related('Room_ID', 'Tenant_ID')

    created = 0
    for contract in contracts:
        # ข้ามถ้ามี invoice ในเดือนนี้แล้ว (ป้องกันการกดปุ่มซ้ำแล้วเบิ้ลบิล)
        if Invoice.objects.filter(
            Contract_ID  = contract,
            Billing_Date__year = bill_date.year,
            Billing_Date__month = bill_date.month
        ).exists():
            continue

        # ดึงข้อมูล utility ที่จดไว้แล้ว
        utility = Utility.objects.filter(
            Room_ID    = contract.Room_ID,
            Bill_Month = bill_month
        ).first()

        # ถ้ายังไม่ได้จดมิเตอร์ → ข้ามห้องนี้ไปก่อน
        if not utility:
            continue

        water_total = utility.Water_Total
        elec_total  = utility.Elec_Total
        grand_total = contract.Rent_Price + water_total + elec_total

        invoice = Invoice.objects.create(
            Contract_ID  = contract,
            Billing_Date = bill_date,
            Due_Date     = due_date,
            Grand_Total  = grand_total,
            Status       = 'รอชำระ',
        )
        MonthlyBill.objects.get_or_create(
            Invoice_ID = invoice,
            defaults={
                'Bill_Month': bill_month,
                'Amount':     contract.Rent_Price,
            }
        )
        # ผูก utility เข้ากับ invoice ที่เพิ่งสร้าง
        utility.Invoice_ID = invoice
        utility.save()

        created += 1

    return created


@login_required
@role_required('ADMIN', 'MANAGER')
def invoice_generate(request):
    """สร้างใบแจ้งหนี้ประจำเดือนด้วยปุ่มmanual (POST เท่านั้น)"""
    import datetime
    today = datetime.date.today()

    if request.method == 'POST':
        created = auto_generate_invoices()
        bill_date = today.replace(day=25)
        month_name = [
            '','มกราคม','กุมภาพันธ์','มีนาคม','เมษายน',
            'พฤษภาคม','มิถุนายน','กรกฎาคม','สิงหาคม',
            'กันยายน','ตุลาคม','พฤศจิกายน','ธันวาคม',
        ][today.month]
        message = f'สร้างใบแจ้งหนี้เดือน{month_name} {today.year} เรียบร้อย {created} ฉบับ'
        if created == 0:
            message = f'ไม่มีใบแจ้งหนี้ที่ต้องสร้างเพิ่ม (ยังไม่จดมิเตอร์ หรือสร้างไปแล้ว)'
        return render(request, 'apartment/invoice/generate_result.html', {
            'created': created,
            'message': message,
            'today':   today,
        })

    # GET → redirect กลับหน้า list
    return redirect('invoice_list')
# ==================== MAINTENANCE ====================

@login_required
@role_required('ADMIN', 'MANAGER')
def maintenance_list(request):
    building = get_user_building(request.user)
    items = Maintenance.objects.select_related('Room_ID').all()
    if building: items = items.filter(Room_ID__Building_No=building)
    items = items.order_by('-Report_Date')
    return render(request, 'apartment/maintenance/list.html', {'items': items})

@login_required
@role_required('ADMIN', 'MANAGER')
def maintenance_create(request):
    form = MaintenanceForm(request.POST or None)
    building = get_user_building(request.user)
    if building:
        form.fields['Room_ID'].queryset = Room.objects.filter(Building_No=building)
    if form.is_valid():
        form.save()
        return redirect('maintenance_list')
    return render(request, 'apartment/maintenance/form.html', {'form': form, 'title': 'เพิ่มรายการแจ้งซ่อม'})

@login_required
@role_required('ADMIN', 'MANAGER')
def maintenance_edit(request, pk):
    item = get_object_or_404(Maintenance, pk=pk)
    form = MaintenanceForm(request.POST or None, instance=item)
    if form.is_valid():
        form.save()
        return redirect('maintenance_list')
    return render(request, 'apartment/maintenance/form.html', {'form': form, 'title': 'อัปเดตการซ่อม'})

@login_required
@role_required('ADMIN')
def maintenance_delete(request, pk):
    item = get_object_or_404(Maintenance, pk=pk)
    if request.method == 'POST':
        item.delete()
        return redirect('maintenance_list')
    return render(request, 'apartment/maintenance/confirm_delete.html', {'object': item, 'title': 'ลบรายการแจ้งซ่อม'})
# ==================== เงินเดือนพนักงาน (Admin only) ====================

@login_required
@role_required('ADMIN')
def salary_list(request):
    employees = EmployeeSalary.objects.all().order_by('Role', 'First_Name')
    total_active = employees.filter(Is_Active=True).aggregate(total=Sum('Monthly_Salary'))['total'] or 0
    return render(request, 'apartment/salary/list.html', {
        'employees': employees,
        'total_active': total_active,
    })


@login_required
@role_required('ADMIN')
def salary_create(request):
    form = EmployeeSalaryForm(request.POST or None)
    if form.is_valid():
        form.save()
        messages.success(request, 'เพิ่มข้อมูลพนักงานเรียบร้อยแล้ว')
        return redirect('salary_list')
    return render(request, 'apartment/salary/form.html', {'form': form, 'title': 'เพิ่มพนักงาน'})


@login_required
@role_required('ADMIN')
def salary_edit(request, pk):
    employee = get_object_or_404(EmployeeSalary, pk=pk)
    form = EmployeeSalaryForm(request.POST or None, instance=employee)
    if form.is_valid():
        form.save()
        messages.success(request, 'แก้ไขข้อมูลพนักงานเรียบร้อยแล้ว')
        return redirect('salary_list')
    return render(request, 'apartment/salary/form.html', {'form': form, 'title': 'แก้ไขข้อมูลพนักงาน', 'employee': employee})


@login_required
@role_required('ADMIN')
def salary_delete(request, pk):
    employee = get_object_or_404(EmployeeSalary, pk=pk)
    if request.method == 'POST':
        employee.delete()
        messages.success(request, 'ลบข้อมูลพนักงานเรียบร้อยแล้ว')
        return redirect('salary_list')
    return render(request, 'apartment/salary/confirm_delete.html', {'object': employee, 'title': 'ลบข้อมูลพนักงาน'})


# ==================== รายงาน ====================

@login_required
@role_required('ADMIN', 'MANAGER')
def invoice_print(request, pk):
    # หน้าพิมพ์ใบแจ้งหนี้ (print-friendly)
    invoice      = get_object_or_404(Invoice, pk=pk)
    monthly_bill = MonthlyBill.objects.filter(Invoice_ID=invoice).first()
    utility      = Utility.objects.filter(Invoice_ID=invoice).first()
    fines        = Fine.objects.filter(Invoice_ID=invoice)
    return render(request, 'apartment/invoice/print.html', {
        'invoice':      invoice,
        'monthly_bill': monthly_bill,
        'utility':      utility,
        'fines':        fines,
    })


@login_required
@role_required('ADMIN', 'MANAGER')
def monthly_summary(request):
    from django.db.models.functions import TruncMonth
    from .middleware import get_user_role
    import json

    building  = get_user_building(request.user)
    user_role = get_user_role(request.user)

    # ── ตัวกรอง GET ──────────────────────────────────────────────────
    filter_year  = request.GET.get('year', '')
    filter_month = request.GET.get('month', '')
    sort_order   = request.GET.get('sort', 'desc')   # asc | desc

    month_names_th = [
        "", "ม.ค.", "ก.พ.", "มี.ค.", "เม.ย.", "พ.ค.", "มิ.ย.",
        "ก.ค.", "ส.ค.", "ก.ย.", "ต.ค.", "พ.ย.", "ธ.ค."
    ]

    # ── ฐาน queryset ─────────────────────────────────────────────────
    invoices_qs = Invoice.objects.select_related(
        'Contract_ID__Tenant_ID', 'Contract_ID__Room_ID'
    ).all()
    if building:
        invoices_qs = invoices_qs.filter(Contract_ID__Room_ID__Building_No=building)
    if filter_year:
        invoices_qs = invoices_qs.filter(Billing_Date__year=filter_year)
    if filter_month:
        invoices_qs = invoices_qs.filter(Billing_Date__month=filter_month)

    order_prefix = '' if sort_order == 'asc' else '-'
    invoices_qs  = invoices_qs.order_by(f'{order_prefix}Billing_Date')

    # ── ข้อมูลรายรับจาก Invoice (paid) ───────────────────────────────
    paid_statuses = ['ชำระแล้ว', 'จ่ายล่าช้า']

    # ค่าเช่า (MonthlyBill) — เฉพาะที่ชำระแล้ว
    rent_qs = MonthlyBill.objects.filter(Invoice_ID__in=invoices_qs.filter(Status__in=paid_statuses))
    if filter_year:
        rent_qs = rent_qs.filter(Invoice_ID__Billing_Date__year=filter_year)
    if filter_month:
        rent_qs = rent_qs.filter(Invoice_ID__Billing_Date__month=filter_month)
    total_rent = rent_qs.aggregate(t=Sum('Amount'))['t'] or 0

    # ค่าปรับ — เฉพาะที่ชำระแล้ว
    fine_qs = Fine.objects.filter(Invoice_ID__in=invoices_qs.filter(Status__in=paid_statuses))
    total_fine = fine_qs.aggregate(t=Sum('Amount'))['t'] or 0

    # ค่าน้ำ/ไฟที่เรียกเก็บจากผู้เช่า (paid)
    util_income_qs = Utility.objects.filter(Invoice_ID__in=invoices_qs.filter(Status__in=paid_statuses))
    total_util_income = util_income_qs.aggregate(
        w=Sum('Water_Total'), e=Sum('Elec_Total')
    )
    total_util_income_amt = (total_util_income['w'] or 0) + (total_util_income['e'] or 0)

    total_income = total_rent + total_fine + total_util_income_amt

    # ── รายจ่าย ───────────────────────────────────────────────────────
    # ค่าน้ำ/ไฟที่ต้องจ่ายให้การไฟฟ้า/ประปา (ทุก Utility ในช่วงที่กรอง)
    util_exp_qs = Utility.objects.filter(Invoice_ID__in=invoices_qs)
    util_exp    = util_exp_qs.aggregate(w=Sum('Water_Total'), e=Sum('Elec_Total'))
    total_utility_exp = (util_exp['w'] or 0) + (util_exp['e'] or 0)

    # ค่าซ่อม — Maintenance ที่ซ่อมเสร็จ กรองตาม Resolved_Date ถ้ากรองเดือน/ปี
    maint_qs = Maintenance.objects.filter(Status='ซ่อมเสร็จ')
    if building:
        maint_qs = maint_qs.filter(Room_ID__Building_No=building)
    if filter_year:
        maint_qs = maint_qs.filter(Resolved_Date__year=filter_year)
    if filter_month:
        maint_qs = maint_qs.filter(Resolved_Date__month=filter_month)
    total_repair = maint_qs.aggregate(t=Sum('Repair_Cost'))['t'] or 0

    # เงินเดือนพนักงาน (ADMIN เห็นเท่านั้น) — รวมทุกเดือนถ้าไม่ได้กรอง
    total_salary = 0
    employees    = []
    if user_role == 'ADMIN':
        employees   = EmployeeSalary.objects.filter(Is_Active=True).order_by('Role')
        month_count = 1
        if filter_year and not filter_month:
            month_count = 12
        total_salary = (employees.aggregate(t=Sum('Monthly_Salary'))['t'] or 0) * month_count

    total_expense = total_utility_exp + total_repair + total_salary
    net_profit    = total_income - total_expense

    # ── สรุปรายเดือน ──────────────────────────────────────────────────
    summary_qs = (
        invoices_qs
        .annotate(month=TruncMonth('Billing_Date'))
        .values('month')
        .annotate(
            total_income=Sum('Grand_Total', filter=Q(Status__in=paid_statuses)),
            count=Count('Invoice_ID'),
            paid=Count('Invoice_ID', filter=Q(Status__in=paid_statuses)),
        )
        .order_by('month' if sort_order == 'asc' else '-month')
    )

    # เตรียมข้อมูลกราฟ (เรียงเก่า→ใหม่เสมอ)
    summary_for_chart = list(summary_qs.order_by('month'))
    chart_labels, chart_income_data, chart_expense_data = [], [], []

    for row in summary_for_chart:
        m   = row['month']
        lbl = f"{month_names_th[m.month]} {m.year + 543}"
        chart_labels.append(lbl)
        chart_income_data.append(float(row['total_income'] or 0))

        # คำนวณรายจ่ายต่อเดือนสำหรับกราฟ
        m_util = Utility.objects.filter(
            Invoice_ID__in=invoices_qs,
            Bill_Month__year=m.year, Bill_Month__month=m.month
        ).aggregate(w=Sum('Water_Total'), e=Sum('Elec_Total'))
        m_repair = Maintenance.objects.filter(
            Status='ซ่อมเสร็จ',
            Resolved_Date__year=m.year, Resolved_Date__month=m.month,
            **({'Room_ID__Building_No': building} if building else {})
        ).aggregate(t=Sum('Repair_Cost'))['t'] or 0
        m_salary = float(total_salary / max(len(summary_for_chart), 1)) if user_role == 'ADMIN' else 0
        m_exp    = float((m_util['w'] or 0) + (m_util['e'] or 0)) + float(m_repair) + m_salary
        chart_expense_data.append(round(m_exp, 2))

    # ปีที่มีข้อมูล (สำหรับ filter dropdown)
    available_years = (
        Invoice.objects.filter(**({'Contract_ID__Room_ID__Building_No': building} if building else {}))
        .dates('Billing_Date', 'year')
    )
    years = [d.year for d in available_years]

    # จำนวน invoice รวม (ไม่ดึง objects ทั้งหมด — เพื่อ performance)
    total_invoice_count = invoices_qs.count()

    return render(request, 'apartment/report/summary.html', {
        'summary':            list(summary_qs),
        'employees':          employees,
        'total_invoice_count': total_invoice_count,
        # totals
        'total_rent':         total_rent,
        'total_fine':         total_fine,
        'total_util_income':  total_util_income_amt,
        'total_income':       total_income,
        'total_utility_exp':  total_utility_exp,
        'total_repair':       total_repair,
        'total_salary':       total_salary,
        'total_expense':      total_expense,
        'net_profit':         net_profit,
        # chart — ส่ง raw list ด้วย เพื่อให้ JS filter เดือนที่ต้องการได้
        'chart_labels_json':       json.dumps(chart_labels),
        'chart_income_json':       json.dumps(chart_income_data),
        'chart_expense_json':      json.dumps(chart_expense_data),
        # filters
        'filter_year':        filter_year,
        'filter_month':       filter_month,
        'sort_order':         sort_order,
        'available_years':    years,
        'month_names_th':     month_names_th,
        'user_role':          user_role,
    })

# ==================== EXPORT EXCEL ====================

@login_required
@role_required('ADMIN', 'MANAGER', 'METER')
def export_summary_excel(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse
    from .middleware import get_user_role
    import io

    building  = get_user_building(request.user)
    user_role = get_user_role(request.user)

    export_type    = request.GET.get('export_type', 'monthly')
    export_year    = request.GET.get('export_year', '')
    compare_year1  = request.GET.get('compare_year1', '')
    compare_month1 = request.GET.get('compare_month1', '')
    compare_year2  = request.GET.get('compare_year2', '')
    compare_month2 = request.GET.get('compare_month2', '')

    inc_income    = request.GET.get('inc_income',    '1') == '1'
    inc_expense   = request.GET.get('inc_expense',   '1') == '1'
    inc_profit    = request.GET.get('inc_profit',    '1') == '1'
    inc_change    = request.GET.get('inc_change',    '1') == '1'
    inc_breakdown = request.GET.get('inc_breakdown', '0') == '1'
    inc_salary    = request.GET.get('inc_salary',    '0') == '1' and user_role == 'ADMIN'

    paid_statuses = ['ชำระแล้ว', 'จ่ายล่าช้า']
    month_names_th = ["", "ม.ค.", "ก.พ.", "มี.ค.", "เม.ย.", "พ.ค.", "มิ.ย.",
                      "ก.ค.", "ส.ค.", "ก.ย.", "ต.ค.", "พ.ย.", "ธ.ค."]
    MONTHS_IN_Q = {1: [1,2,3], 2: [4,5,6], 3: [7,8,9], 4: [10,11,12]}

    def get_period_data(year=None, month=None, quarter=None):
        inv_qs = Invoice.objects.all()
        if building:
            inv_qs = inv_qs.filter(Contract_ID__Room_ID__Building_No=building)
        if year:
            inv_qs = inv_qs.filter(Billing_Date__year=int(year))
        if month:
            inv_qs = inv_qs.filter(Billing_Date__month=int(month))
        if quarter:
            inv_qs = inv_qs.filter(Billing_Date__month__in=MONTHS_IN_Q[quarter])

        paid_inv = inv_qs.filter(Status__in=paid_statuses)
        rent     = MonthlyBill.objects.filter(Invoice_ID__in=paid_inv).aggregate(t=Sum('Amount'))['t'] or 0
        fine     = Fine.objects.filter(Invoice_ID__in=paid_inv).aggregate(t=Sum('Amount'))['t'] or 0
        ui       = Utility.objects.filter(Invoice_ID__in=paid_inv).aggregate(w=Sum('Water_Total'), e=Sum('Elec_Total'))
        util_inc = (ui['w'] or 0) + (ui['e'] or 0)
        income   = rent + fine + util_inc

        ue       = Utility.objects.filter(Invoice_ID__in=inv_qs).aggregate(w=Sum('Water_Total'), e=Sum('Elec_Total'))
        util_exp = (ue['w'] or 0) + (ue['e'] or 0)

        maint_qs = Maintenance.objects.filter(Status='ซ่อมเสร็จ')
        if building:
            maint_qs = maint_qs.filter(Room_ID__Building_No=building)
        if year:
            maint_qs = maint_qs.filter(Resolved_Date__year=int(year))
        if month:
            maint_qs = maint_qs.filter(Resolved_Date__month=int(month))
        if quarter:
            maint_qs = maint_qs.filter(Resolved_Date__month__in=MONTHS_IN_Q[quarter])
        repair = maint_qs.aggregate(t=Sum('Repair_Cost'))['t'] or 0

        salary = 0
        if user_role == 'ADMIN':
            mul = 1 if month else (3 if quarter else 12)
            salary = (EmployeeSalary.objects.filter(Is_Active=True).aggregate(t=Sum('Monthly_Salary'))['t'] or 0) * mul

        expense = util_exp + repair + salary
        return {
            'income': float(income), 'expense': float(expense),
            'profit': float(income - expense),
            'rent': float(rent), 'fine': float(fine), 'util_income': float(util_inc),
            'util_exp': float(util_exp), 'repair': float(repair), 'salary': float(salary),
        }

    def calc_pct(old, new):
        if old == 0:
            return None
        return (new - old) / abs(old)

    # ── Style helpers ──
    C_DARK_PURPLE = 'FF4A00E0'
    C_LIGHT_PURPLE = 'FFede9fe'
    C_GREEN  = 'FF059669'
    C_RED    = 'FFdc2626'
    C_LGREEN = 'FFd1fae5'
    C_LRED   = 'FFfee2e2'
    C_GRAY   = 'FF6b7280'
    C_WHITE  = 'FFFFFFFF'
    C_ALT    = 'FFF9FAFB'

    thin = Border(
        left=Side(style='thin', color='FFE5E7EB'), right=Side(style='thin', color='FFE5E7EB'),
        top=Side(style='thin', color='FFE5E7EB'),  bottom=Side(style='thin', color='FFE5E7EB'),
    )
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    right  = Alignment(horizontal='right',  vertical='center')
    left   = Alignment(horizontal='left',   vertical='center')

    def hfont(bold=True, size=10, color=C_WHITE): return Font(name='Calibri', bold=bold, size=size, color=color)
    def fill(color): return PatternFill('solid', fgColor=color)
    def sc(cell, fnt=None, fl=None, aln=None, fmt=None):
        cell.border = thin
        if fnt: cell.font = fnt
        if fl:  cell.fill = fl
        if aln: cell.alignment = aln
        if fmt: cell.number_format = fmt

    NUM  = '#,##0.00'
    PCT  = '+0.00%;-0.00%;0.00%'

    wb = openpyxl.Workbook()

    # ── Available years ──
    base_inv = Invoice.objects.all()
    if building:
        base_inv = base_inv.filter(Contract_ID__Room_ID__Building_No=building)
    all_years = sorted(base_inv.values_list('Billing_Date__year', flat=True).distinct())

    def make_title_row(ws, text, ncols):
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
        c = ws.cell(row=1, column=1, value=text)
        c.font = hfont(size=14); c.fill = fill(C_DARK_PURPLE); c.alignment = center; c.border = thin
        ws.row_dimensions[1].height = 34
        import datetime
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
        s = ws.cell(row=2, column=1, value=f'สร้างเมื่อ: {datetime.date.today().strftime("%d/%m/") + str(datetime.date.today().year + 543)}')
        s.font = Font(name='Calibri', size=9, color=C_GRAY); s.fill = fill(C_LIGHT_PURPLE); s.alignment = center

    def header_row(ws, row_no, headers):
        for i, h in enumerate(headers, 1):
            c = ws.cell(row=row_no, column=i, value=h)
            sc(c, fnt=hfont(), fl=fill(C_DARK_PURPLE), aln=center)
        ws.row_dimensions[row_no].height = 22

    # ─────────────────────────────────────────────────────────────────
    # Monthly / Quarterly / Annual   →  one sheet
    # ─────────────────────────────────────────────────────────────────
    if export_type in ('monthly', 'quarterly', 'annual'):
        ws = wb.active
        title_map = {
            'monthly':  f'รายงานรายเดือน{" ปี " + str(int(export_year)+543) if export_year else ""}',
            'quarterly':f'รายงานรายไตรมาส{" ปี " + str(int(export_year)+543) if export_year else ""}',
            'annual':   'รายงานสรุปรายปี',
        }
        ws.title = {'monthly':'รายเดือน','quarterly':'รายไตรมาส','annual':'รายปี'}[export_type]

        # Build header list
        headers = ['งวด']
        col_map = {}
        ci = 2
        if inc_income:    headers.append('รายรับ (฿)');     col_map['income']     = ci; ci += 1
        if inc_expense:   headers.append('รายจ่าย (฿)');    col_map['expense']    = ci; ci += 1
        if inc_profit:    headers.append('กำไรสุทธิ (฿)');  col_map['profit']     = ci; ci += 1
        if inc_change and inc_income:
            headers.append('% เปลี่ยนแปลง รายรับ'); col_map['inc_pct'] = ci; ci += 1
        if inc_change and inc_profit:
            headers.append('% เปลี่ยนแปลง กำไร');  col_map['prf_pct'] = ci; ci += 1
        if inc_breakdown:
            for h, k in [('ค่าเช่า','rent'),('ค่าน้ำ/ไฟ(เก็บ)','util_income'),
                         ('ค่าปรับ','fine'),('ค่าน้ำ/ไฟ(จ่าย)','util_exp'),('ค่าซ่อม','repair')]:
                headers.append(h); col_map[k] = ci; ci += 1
            if inc_salary:
                headers.append('เงินเดือน'); col_map['salary'] = ci; ci += 1

        ncols = len(headers)
        make_title_row(ws, title_map[export_type], ncols)
        header_row(ws, 4, headers)
        ws.column_dimensions['A'].width = 20
        for c_i in range(2, ci):
            ws.column_dimensions[get_column_letter(c_i)].width = 20

        # Build periods list
        if export_type == 'monthly':
            years_use = [int(export_year)] if export_year else all_years
            periods = [{'label': f'{month_names_th[m]} {y+543}', 'year': y, 'month': m}
                       for y in years_use for m in range(1, 13)]
        elif export_type == 'quarterly':
            years_use = [int(export_year)] if export_year else all_years
            periods = [{'label': f'Q{q} ปี {y+543}', 'year': y, 'quarter': q}
                       for y in years_use for q in range(1, 5)]
        else:
            periods = [{'label': f'ปี {y+543}', 'year': y} for y in all_years]

        prev_data = None
        row_i = 5
        tot_inc = tot_exp = tot_prf = 0

        for pi, period in enumerate(periods):
            data = get_period_data(year=period.get('year'), month=period.get('month'), quarter=period.get('quarter'))
            if export_type != 'monthly' and data['income'] == 0 and data['expense'] == 0:
                prev_data = data; continue

            row_fill = fill(C_ALT) if pi % 2 == 0 else fill(C_WHITE)

            c = ws.cell(row=row_i, column=1, value=period['label'])
            sc(c, fnt=hfont(bold=True, color='FF2b2d42'), fl=row_fill, aln=left)

            if inc_income:
                c = ws.cell(row=row_i, column=col_map['income'], value=data['income'])
                sc(c, fnt=hfont(color=C_GREEN), fl=fill(C_LGREEN), aln=right, fmt=NUM)
            if inc_expense:
                c = ws.cell(row=row_i, column=col_map['expense'], value=data['expense'])
                sc(c, fnt=hfont(color=C_RED), fl=fill(C_LRED), aln=right, fmt=NUM)
            if inc_profit:
                pf = C_GREEN if data['profit'] >= 0 else C_RED
                pfl = fill(C_LGREEN) if data['profit'] >= 0 else fill(C_LRED)
                c = ws.cell(row=row_i, column=col_map['profit'], value=data['profit'])
                sc(c, fnt=hfont(color=pf), fl=pfl, aln=right, fmt=NUM)
            if inc_change and 'inc_pct' in col_map and prev_data:
                pct = calc_pct(prev_data['income'], data['income'])
                if pct is not None:
                    c = ws.cell(row=row_i, column=col_map['inc_pct'], value=pct)
                    sc(c, fnt=hfont(color=C_GREEN if pct >= 0 else C_RED), fl=row_fill, aln=center, fmt=PCT)
            if inc_change and 'prf_pct' in col_map and prev_data:
                pct = calc_pct(prev_data['profit'], data['profit'])
                if pct is not None:
                    c = ws.cell(row=row_i, column=col_map['prf_pct'], value=pct)
                    sc(c, fnt=hfont(color=C_GREEN if pct >= 0 else C_RED), fl=row_fill, aln=center, fmt=PCT)
            if inc_breakdown:
                for key in ['rent','util_income','fine','util_exp','repair']:
                    if key in col_map:
                        c = ws.cell(row=row_i, column=col_map[key], value=data[key])
                        sc(c, fnt=hfont(bold=False, color='FF374151'), fl=row_fill, aln=right, fmt=NUM)
                if inc_salary and 'salary' in col_map:
                    c = ws.cell(row=row_i, column=col_map['salary'], value=data['salary'])
                    sc(c, fnt=hfont(bold=False, color='FF374151'), fl=row_fill, aln=right, fmt=NUM)

            tot_inc += data['income']; tot_exp += data['expense']; tot_prf += data['profit']
            prev_data = data; row_i += 1

        # Total row
        tf = fill(C_LIGHT_PURPLE)
        c = ws.cell(row=row_i, column=1, value='รวมทั้งหมด')
        sc(c, fnt=hfont(bold=True, size=11, color='FF4A00E0'), fl=tf, aln=left)
        if inc_income:
            c = ws.cell(row=row_i, column=col_map['income'], value=tot_inc)
            sc(c, fnt=hfont(bold=True, size=11, color=C_GREEN), fl=tf, aln=right, fmt=NUM)
        if inc_expense:
            c = ws.cell(row=row_i, column=col_map['expense'], value=tot_exp)
            sc(c, fnt=hfont(bold=True, size=11, color=C_RED), fl=tf, aln=right, fmt=NUM)
        if inc_profit:
            c = ws.cell(row=row_i, column=col_map['profit'], value=tot_prf)
            pf = C_GREEN if tot_prf >= 0 else C_RED
            sc(c, fnt=hfont(bold=True, size=11, color=pf), fl=tf, aln=right, fmt=NUM)
        ws.row_dimensions[row_i].height = 24
        ws.freeze_panes = 'A5'

    # ─────────────────────────────────────────────────────────────────
    # Comparison  →  one sheet
    # ─────────────────────────────────────────────────────────────────
    elif export_type == 'comparison':
        ws = wb.active
        ws.title = 'เปรียบเทียบ 2 ช่วง'

        def period_label(year, month):
            if year and month:
                return f'{month_names_th[int(month)]} {int(year)+543}'
            elif year:
                return f'ปี {int(year)+543}'
            return 'ทุกช่วง'

        label1 = period_label(compare_year1, compare_month1)
        label2 = period_label(compare_year2, compare_month2)
        data1  = get_period_data(year=compare_year1 or None, month=compare_month1 or None)
        data2  = get_period_data(year=compare_year2 or None, month=compare_month2 or None)

        headers = ['รายการ', label1, label2, 'เปลี่ยนแปลง (฿)', '% เปลี่ยนแปลง', 'แนวโน้ม']
        ncols = 6
        make_title_row(ws, f'รายงานเปรียบเทียบ: {label1} vs {label2}', ncols)
        header_row(ws, 4, headers)

        ws.column_dimensions['A'].width = 26
        for ltr in 'BCDEF':
            ws.column_dimensions[ltr].width = 22

        rows_def = [
            ('รายรับรวม',           data1['income'],      data2['income'],      True,  True),
            ('  └ ค่าเช่า',         data1['rent'],        data2['rent'],        False, True),
            ('  └ ค่าน้ำ/ไฟ (เก็บ)',data1['util_income'], data2['util_income'], False, True),
            ('  └ ค่าปรับ',         data1['fine'],        data2['fine'],        False, True),
            ('รายจ่ายรวม',          data1['expense'],     data2['expense'],     True,  False),
            ('  └ ค่าน้ำ/ไฟ (จ่าย)',data1['util_exp'],   data2['util_exp'],    False, False),
            ('  └ ค่าซ่อมบำรุง',    data1['repair'],      data2['repair'],      False, False),
        ]
        if user_role == 'ADMIN':
            rows_def.append(('  └ เงินเดือน', data1['salary'], data2['salary'], False, False))
        rows_def.append(('กำไรสุทธิ', data1['profit'], data2['profit'], True, True))

        for ri, (label, v1, v2, is_section, is_inc) in enumerate(rows_def, 5):
            row_fl = fill(C_LIGHT_PURPLE) if is_section else (fill(C_LGREEN) if is_inc else fill(C_LRED))
            fnt_sz = 11 if is_section else 10

            c = ws.cell(row=ri, column=1, value=label)
            sc(c, fnt=hfont(bold=is_section, size=fnt_sz, color='FF2b2d42'), fl=row_fl, aln=left)

            for col_no, val in [(2, v1), (3, v2)]:
                c = ws.cell(row=ri, column=col_no, value=val)
                sc(c, fnt=hfont(bold=is_section, size=fnt_sz, color='FF2b2d42'), fl=row_fl, aln=right, fmt=NUM)

            diff = v2 - v1
            c = ws.cell(row=ri, column=4, value=diff)
            sc(c, fnt=hfont(bold=is_section, size=fnt_sz, color=C_GREEN if diff >= 0 else C_RED),
               fl=row_fl, aln=right, fmt='+#,##0.00;-#,##0.00;0.00')

            pct = calc_pct(v1, v2)
            if pct is not None:
                c = ws.cell(row=ri, column=5, value=pct)
                sc(c, fnt=hfont(bold=is_section, size=fnt_sz, color=C_GREEN if pct >= 0 else C_RED),
                   fl=row_fl, aln=center, fmt=PCT)

            trend = '▲ เพิ่มขึ้น' if diff > 0 else ('▼ ลดลง' if diff < 0 else '─ คงที่')
            c = ws.cell(row=ri, column=6, value=trend)
            sc(c, fnt=hfont(bold=False, size=fnt_sz, color=C_GREEN if diff > 0 else (C_RED if diff < 0 else C_GRAY)),
               fl=row_fl, aln=center)
            ws.row_dimensions[ri].height = 20

        ws.freeze_panes = 'A5'

    # ── Build response ──
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    fname_map = {
        'monthly':    f'report_monthly{"_" + export_year if export_year else ""}.xlsx',
        'comparison': f'report_compare.xlsx',
        'quarterly':  f'report_quarterly{"_" + export_year if export_year else ""}.xlsx',
        'annual':     'report_annual.xlsx',
    }
    fname = fname_map.get(export_type, 'report.xlsx')

    from django.http import HttpResponse
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{fname}"'
    return response


# ==================== API ====================

from django.http import JsonResponse


@login_required
@role_required('ADMIN', 'MANAGER')
def api_invoices_by_month(request):
    """AJAX: คืน JSON รายการ Invoice ของเดือน/ปีที่ระบุ"""
    year  = request.GET.get('year')
    month = request.GET.get('month')
    if not year or not month:
        return JsonResponse({'error': 'year and month required'}, status=400)

    building = get_user_building(request.user)
    qs = Invoice.objects.select_related(
        'Contract_ID__Tenant_ID', 'Contract_ID__Room_ID'
    ).filter(Billing_Date__year=year, Billing_Date__month=month)
    if building:
        qs = qs.filter(Contract_ID__Room_ID__Building_No=building)
    qs = qs.order_by('Billing_Date', 'Invoice_ID')

    data = []
    for inv in qs:
        data.append({
            'id':        inv.Invoice_ID,
            'tenant':    f"{inv.Contract_ID.Tenant_ID.First_Name} {inv.Contract_ID.Tenant_ID.Last_Name}",
            'room':      inv.Contract_ID.Room_ID.Room_Number,
            'date':      inv.Billing_Date.strftime('%d/%m/%Y'),
            'total':     float(inv.Grand_Total),
            'status':    inv.Status,
            'print_url': f"/invoices/{inv.Invoice_ID}/print/",
        })
    return JsonResponse({'invoices': data})


@login_required
def api_rooms_available(request):
    """JSON API สำหรับ cascading filter อาคาร → ชั้น → ห้อง (เฉพาะห้องว่างที่ยังไม่ถูกจอง)"""
    rooms = Room.objects.filter(Status='ว่าง', Status_Flag='ปกติ')
    building = request.GET.get('building')
    floor = request.GET.get('floor')

    if building:
        rooms = rooms.filter(Building_No=building)
    if floor:
        rooms = rooms.filter(Floor=floor)

    # ถ้าขอแค่ buildings
    if request.GET.get('type') == 'buildings':
        user_building = get_user_building(request.user)
        bld_qs = Room.objects.filter(Status='ว่าง', Status_Flag='ปกติ')
        if user_building:
            bld_qs = bld_qs.filter(Building_No=user_building)
        buildings = bld_qs.values_list('Building_No', flat=True).distinct().order_by('Building_No')
        return JsonResponse({'buildings': list(buildings)})

    # ถ้าขอ floors ของ building
    if request.GET.get('type') == 'floors' and building:
        floors = rooms.values_list('Floor', flat=True).distinct().order_by('Floor')
        return JsonResponse({'floors': list(floors)})

    # ถ้าขอ rooms
    data = list(rooms.values('Room_ID', 'Room_Number', 'Building_No', 'Floor').order_by('Room_Number'))
    return JsonResponse({'rooms': data})

@login_required
def api_utility_latest(request):
    """JSON API สำหรับดึงข้อมูลมิเตอร์ล่าสุดของสัญญา เพื่อ auto-fill ในฟอร์มออกใบแจ้งหนี้"""
    contract_id = request.GET.get('contract_id')
    if not contract_id:
        return JsonResponse({'error': 'Missing contract_id'}, status=400)
    
    contract = Contract.objects.filter(pk=contract_id).first()
    if not contract:
        return JsonResponse({'error': 'Invalid contract_id'}, status=400)

    import datetime
    today = datetime.date.today()
    bill_month = today.replace(day=1)
    
    # ดึงค่า Utility ล่าสุดของห้องนี้
    utility = Utility.objects.filter(Room_ID=contract.Room_ID).order_by('-Bill_Month').first()
    
    data = {
        'Water_Cost_Unit': contract.Water_Cost_Unit,
        'Elec_Cost_Unit': contract.Elec_Cost_Unit,
        'Bill_Month': bill_month.strftime('%Y-%m-%d'),
    }

    if utility:
        data.update({
            'Water_Unit_Used': utility.Water_Unit_Used,
            'Elec_Unit_Used': utility.Elec_Unit_Used,
        })
    else:
        data.update({
            'Water_Unit_Used': 0,
            'Elec_Unit_Used': 0,
        })
    return JsonResponse(data)

@login_required
def api_room_meter_latest(request):
    """JSON API สำหรับดึงข้อมูลมิเตอร์ล่าสุดของห้อง เพื่อ auto-fill ในฟอร์มสร้างสัญญา"""
    room_id = request.GET.get('room_id')
    if not room_id:
        return JsonResponse({'error': 'Missing room_id'}, status=400)
    
    # ดึง Utility ล่าสุดของห้องนี้
    latest_u = Utility.objects.filter(Room_ID_id=room_id).order_by('-Bill_Month').first()
    
    if latest_u:
        return JsonResponse({
            'water_start': latest_u.Water_Unit_After,
            'elec_start': latest_u.Elec_Unit_After,
        })
    else:
        # ถ้ายังไม่มีประวัติ Utility เลย ให้ลองดึงจากสัญญาเก่าล่าสุด (ถ้ามี)
        latest_c = Contract.objects.filter(Room_ID_id=room_id).order_by('-Contract_ID').first()
        if latest_c:
            return JsonResponse({
                'water_start': latest_c.Water_Meter_Start,
                'elec_start': latest_c.Elec_Meter_Start,
            })
    
    return JsonResponse({'water_start': 0, 'elec_start': 0})

# ==================== BOOKING ====================

@login_required
@role_required('ADMIN', 'MANAGER')
def booking_list(request):
    building = get_user_building(request.user)
    bookings = Booking.objects.select_related('Room_ID').filter(Status='รอยืนยัน')
    if building: bookings = bookings.filter(Room_ID__Building_No=building)
    bookings = bookings.order_by('-Booking_Date')
    return render(request, 'apartment/booking/list.html', {'bookings': bookings})


@login_required
@role_required('ADMIN', 'MANAGER')
def booking_create(request, room_pk=None):
    initial = {}
    if room_pk:
        room = get_object_or_404(Room, pk=room_pk)
        initial['Room_ID'] = room

    form = BookingForm(request.POST or None, initial=initial)
    # กรองเฉพาะห้องว่างที่ยังไม่ถูกจอง และ filter ตาม building ของ user
    available_rooms = Room.objects.filter(Status='ว่าง', Status_Flag='ปกติ')
    building = get_user_building(request.user)
    if building:
        available_rooms = available_rooms.filter(Building_No=building)
    form.fields['Room_ID'].queryset = available_rooms

    if form.is_valid():
        booking = form.save(commit=False)
        booking.Status = 'รอยืนยัน'
        booking.save()
        # อัปเดตสถานะห้องเป็น จอง
        room             = booking.Room_ID
        room.Status_Flag = 'จอง'
        room.save()
        return redirect('booking_list')

    return render(request, 'apartment/booking/form.html', {
        'form':  form,
        'title': 'บันทึกการจองห้อง',
    })


@login_required
@role_required('ADMIN', 'MANAGER')
def booking_cancel(request, pk):
    booking = get_object_or_404(Booking, pk=pk)
    if request.method == 'POST':
        # คืนสถานะห้องเป็นปกติ
        room             = booking.Room_ID
        room.Status_Flag = 'ปกติ'
        room.save()
        booking.Status = 'ยกเลิก'
        booking.save()
        return redirect('booking_list')
    return render(request, 'apartment/booking/confirm_cancel.html', {'booking': booking})


@login_required
@role_required('ADMIN', 'MANAGER')
def booking_confirm(request, pk):
    booking = get_object_or_404(Booking, pk=pk)

    if request.method == 'POST':
        contract_form = ContractForm(request.POST)
        contract_form.fields['Room_ID'].queryset = Room.objects.filter(
            Room_ID=booking.Room_ID.Room_ID
        )
        # ลบ Tenant_ID ออกจาก required เพราะเราจะสร้างเองใน code
        contract_form.fields['Tenant_ID'].required = False

        if contract_form.is_valid():
            room = booking.Room_ID
            # ป้องกัน race condition: ห้องนี้มีสัญญา active อยู่แล้วหรือยัง
            if Contract.objects.filter(Room_ID=room, Status='ใช้งาน').exists():
                contract_form.add_error(None, 'ห้องนี้มีสัญญาใช้งานอยู่แล้ว ไม่สามารถยืนยันการจองซ้ำได้')
            else:
                # 1. สร้าง Tenant จากข้อมูลการจอง
                tenant = Tenant.objects.create(
                    First_Name = booking.First_Name,
                    Last_Name  = booking.Last_Name,
                    ID_Card    = booking.ID_Card,
                    Phone      = booking.Phone,
                    Email      = booking.Email   or '',
                    Line_ID    = booking.Line_ID or '',
                    Address    = booking.Address or '',
                )
                # 2. สร้าง Contract แล้วผูก Tenant
                contract           = contract_form.save(commit=False)
                contract.Tenant_ID = tenant
                contract.save()
                # 3. อัปเดตสถานะห้อง
                room.Status      = 'มีผู้เช่า'
                room.Status_Flag = 'ปกติ'
                room.save()
                # 4. ยกเลิกการจองอื่นๆ ของห้องเดียวกันที่ยังค้างอยู่ (ถ้ามี)
                Booking.objects.filter(Room_ID=room, Status='รอยืนยัน').exclude(pk=booking.pk).update(Status='ยกเลิก')
                # 5. ปิดการจองนี้
                booking.Status = 'ยืนยันแล้ว'
                booking.save()
                return redirect('contract_print', pk=contract.Contract_ID)
    else:
        room = booking.Room_ID
        meter_initial = {}
        latest_u = Utility.objects.filter(Room_ID=room).order_by('-Bill_Month').first()
        if latest_u:
            meter_initial['Water_Meter_Start'] = latest_u.Water_Unit_After
            meter_initial['Elec_Meter_Start']  = latest_u.Elec_Unit_After
        else:
            latest_c = Contract.objects.filter(Room_ID=room).order_by('-Contract_ID').first()
            if latest_c:
                meter_initial['Water_Meter_Start'] = latest_c.Water_Meter_Start
                meter_initial['Elec_Meter_Start']  = latest_c.Elec_Meter_Start

        contract_form = ContractForm(initial={
            'Room_ID':          booking.Room_ID,
            'Rent_Price':       4000,
            'Deposit':          4000,
            'Deposit_Advance':  2000,
            'Water_Cost_Unit':  18,
            'Elec_Cost_Unit':   8,
            'Status':           'ใช้งาน',
            **meter_initial,
        })
        contract_form.fields['Room_ID'].queryset = Room.objects.filter(
            Room_ID=booking.Room_ID.Room_ID
        )
        # ซ่อน Tenant_ID ออกจากฟอร์มเลย เพราะไม่ต้องให้ user เลือก
        contract_form.fields['Tenant_ID'].required = False
        contract_form.fields['Tenant_ID'].widget   = contract_form.fields['Tenant_ID'].hidden_widget()
        # ซ่อน Status ด้วย เพราะตั้งค่าเป็น 'ใช้งาน' อัตโนมัติ
        contract_form.fields['Status'].widget = contract_form.fields['Status'].hidden_widget()

    return render(request, 'apartment/booking/confirm.html', {
        'booking':       booking,
        'contract_form': contract_form,
    })

# ==================== METER ====================

@login_required
@role_required('ADMIN', 'MANAGER', 'METER')
def meter_index(request):
    from .middleware import get_user_role
    if get_user_role(request.user) == 'METER':
        return redirect('meter_input')
    import datetime
    today = datetime.date.today()

    # รับเดือน/ปีที่เลือก (default = เดือนปัจจุบัน)
    month = int(request.GET.get('month', today.month))
    year  = int(request.GET.get('year',  today.year))
    bill_month = datetime.date(year, month, 1)

    # เดือนก่อนหน้า
    if month == 1:
        prev_month = datetime.date(year - 1, 12, 1)
    else:
        prev_month = datetime.date(year, month - 1, 1)

    # ดึงห้องที่มีผู้เช่าเท่านั้น จัดกลุ่มตามอาคาร/ชั้น (filter ตาม building ของ user)
    building = get_user_building(request.user)
    rooms = Room.objects.filter(Status='มีผู้เช่า')
    if building:
        rooms = rooms.filter(Building_No=building)
    rooms = rooms.order_by('Building_No', 'Floor', 'Room_Number')

    # ดึง utility เดือนก่อนหน้า → ใช้เป็น "เลขก่อนหน้า"
    prev_utilities = Utility.objects.filter(Bill_Month=prev_month)
    prev_map = {u.Room_ID_id: u for u in prev_utilities}

    # ดึง utility เดือนปัจจุบัน (ถ้าบันทึกไปแล้ว)
    curr_utilities = Utility.objects.filter(Bill_Month=bill_month)
    curr_map = {u.Room_ID_id: u for u in curr_utilities}

    # ดึง contract เพื่อรู้ค่าน้ำ/ไฟต่อหน่วย
    contracts = Contract.objects.filter(
        Status='ใช้งาน'
    ).select_related('Room_ID')
    contract_map = {c.Room_ID_id: c for c in contracts}

    # จัดกลุ่มห้องตามอาคาร→ชั้น
    from itertools import groupby
    buildings = {}
    for room in rooms:
        b = room.Building_No
        f = room.Floor
        if b not in buildings:
            buildings[b] = {}
        if f not in buildings[b]:
            buildings[b][f] = []

        curr_u    = curr_map.get(room.Room_ID)
        contract  = contract_map.get(room.Room_ID)

        latest_u = Utility.objects.filter(
            Room_ID=room
        ).exclude(
            Bill_Month=bill_month
        ).order_by('-Bill_Month').first()

        buildings[b][f].append({
            'room':     room,
            'contract': contract,
            'prev_u':   latest_u,
            'curr_u':   curr_u,
            'water_prev': latest_u.Water_Unit_After if latest_u else (contract.Water_Meter_Start if contract else 0),
            'elec_prev':  latest_u.Elec_Unit_After if latest_u else (contract.Elec_Meter_Start if contract else 0),
        })

    # dropdown ตัวเลือก
    months_th = [
        (1,'มกราคม'),(2,'กุมภาพันธ์'),(3,'มีนาคม'),(4,'เมษายน'),
        (5,'พฤษภาคม'),(6,'มิถุนายน'),(7,'กรกฎาคม'),(8,'สิงหาคม'),
        (9,'กันยายน'),(10,'ตุลาคม'),(11,'พฤศจิกายน'),(12,'ธันวาคม'),
    ]
    years     = list(range(today.year - 2, today.year + 2))

    return render(request, 'apartment/meter/index.html', {
        'buildings':  buildings,
        'month':      month,
        'year':       year,
        'bill_month': bill_month,
        'today':      today,
        'months_th':  months_th,
        'years':      years,
    })


@login_required
@role_required('ADMIN', 'MANAGER', 'METER')
def meter_save(request):
    import datetime
    if request.method != 'POST':
        return redirect('meter_index')

    month      = int(request.POST.get('month'))
    year       = int(request.POST.get('year'))
    record_date = request.POST.get('record_date')
    bill_month = datetime.date(year, month, 1)

    rooms = Room.objects.filter(Status='มีผู้เช่า')
    contract_map = {
        c.Room_ID_id: c
        for c in Contract.objects.filter(Status='ใช้งาน')
    }

    saved = 0
    for room in rooms:
        water_after_key = f"water_after_{room.Room_ID}"
        elec_after_key  = f"elec_after_{room.Room_ID}"

        water_after = request.POST.get(water_after_key, '').strip()
        elec_after  = request.POST.get(elec_after_key,  '').strip()

        # ข้ามถ้าไม่ได้กรอก
        if not water_after or not elec_after:
            continue

        contract = contract_map.get(room.Room_ID)
        if not contract:
            continue

        # ดึงเลขก่อนหน้า
        if month == 1:
            prev_month = datetime.date(year - 1, 12, 1)
        else:
            prev_month = datetime.date(year, month - 1, 1)

        prev_u = Utility.objects.filter(
            Room_ID=room, Bill_Month=prev_month
        ).first()

        from decimal import Decimal

        water_before = prev_u.Water_Unit_After if prev_u else contract.Water_Meter_Start
        elec_before  = prev_u.Elec_Unit_After if prev_u else contract.Elec_Meter_Start

        water_after_d = Decimal(str(water_after))
        elec_after_d  = Decimal(str(elec_after))
        
        water_used    = water_after_d - water_before
        if water_used < Decimal('0'):
            continue
        elec_used     = elec_after_d  - elec_before

        water_total = water_used * Decimal(str(contract.Water_Cost_Unit))
        elec_total  = elec_used  * Decimal(str(contract.Elec_Cost_Unit))

        # สร้าง Invoice สำหรับเดือนนี้ก่อน (ถ้ายังไม่มี)
        invoice, created = Invoice.objects.get_or_create(
            Contract_ID  = contract,
            Billing_Date = bill_month,
            defaults={
                'Due_Date':    bill_month.replace(day=15),
                'Grand_Total': 0,
                'Status':      'รอชำระ',
            }
        )

        # บันทึก/อัปเดต Utility
        Utility.objects.update_or_create(
            Invoice_ID = invoice,
            Room_ID    = room,
            defaults={
                'Bill_Month':        bill_month,
                'Water_Unit_Before': water_before,
                'Water_Unit_After':  water_after_d,
                'Water_Unit_Used':   water_used,
                'Elec_Unit_Before':  elec_before,
                'Elec_Unit_After':   elec_after_d,
                'Elec_Unit_Used':    elec_used,
                'Water_Cost_Unit':   contract.Water_Cost_Unit,
                'Elec_Cost_Unit':    contract.Elec_Cost_Unit,
                'Water_Total':       water_total,
                'Elec_Total':        elec_total,
            }
        )

        # บันทึก/อัปเดต MonthlyBill
        MonthlyBill.objects.update_or_create(
            Invoice_ID = invoice,
            defaults={
                'Bill_Month': bill_month,
                'Amount':     contract.Rent_Price,
            }
        )

        # อัปเดต Grand_Total ของ Invoice
        fine_total = Fine.objects.filter(
            Invoice_ID=invoice
        ).aggregate(t=Sum('Amount'))['t'] or 0

        invoice.Grand_Total = contract.Rent_Price + water_total + elec_total + fine_total
        invoice.save()
        saved += 1

    return redirect(f"/meter/?month={month}&year={year}&saved={saved}")

@login_required
@role_required('ADMIN', 'MANAGER', 'METER')
def meter_input(request):
    import datetime
    today = datetime.date.today()
    month = today.month
    year  = today.year
    bill_month = datetime.date(year, month, 1)

    if month == 1:
        prev_month = datetime.date(year - 1, 12, 1)
    else:
        prev_month = datetime.date(year, month - 1, 1)

    building = get_user_building(request.user)
    rooms = Room.objects.filter(Status='มีผู้เช่า')
    if building: rooms = rooms.filter(Building_No=building)
    rooms = rooms.order_by('Building_No', 'Floor', 'Room_Number')
    contract_map = {c.Room_ID_id: c for c in Contract.objects.filter(Status='ใช้งาน')}
    prev_map     = {u.Room_ID_id: u for u in Utility.objects.filter(Bill_Month=prev_month)}
    curr_map     = {u.Room_ID_id: u for u in Utility.objects.filter(Bill_Month=bill_month)}

    # จัดกลุ่มตามอาคาร
    buildings = {}
    for room in rooms:
        b = room.Building_No
        if b not in buildings:
            buildings[b] = []
        contract = contract_map.get(room.Room_ID)
        curr_u   = curr_map.get(room.Room_ID)
        latest_u = Utility.objects.filter(
            Room_ID=room
        ).exclude(
            Bill_Month=bill_month
        ).order_by('-Bill_Month').first()
        buildings[b].append({
            'room':       room,
            'curr_u':     curr_u,
            'water_prev': latest_u.Water_Unit_After if latest_u else (contract.Water_Meter_Start if contract else 0),
            'elec_prev':  latest_u.Elec_Unit_After if latest_u else (contract.Elec_Meter_Start if contract else 0),
        })

    if request.method == 'POST':
        return redirect('meter_save_input')

    # ดึงรายการตึก/ชั้นสำหรับ filter dropdown
    buildings_list = sorted(buildings.keys())
    floors_map = {}
    for b, room_list in buildings.items():
        floors_map[b] = sorted(set(item['room'].Floor for item in room_list))

    import json
    return render(request, 'apartment/meter/input.html', {
        'buildings':      buildings,
        'buildings_list': buildings_list,
        'floors_map_json': json.dumps(floors_map),
        'month':          month,
        'year':           year,
        'today':          today,
    })

# ==================== ROOM ACTIONS ====================

@login_required
@role_required('ADMIN', 'MANAGER')
def room_action_moveout(request, pk):
    room     = get_object_or_404(Room, pk=pk)
    contract = Contract.objects.filter(Room_ID=room, Status='ใช้งาน').first()

    # คำนวณยอดค้างชำระรวม (รอชำระ + เกินกำหนด)
    unpaid_invoices = Invoice.objects.filter(
        Contract_ID=contract, Status__in=['รอชำระ', 'เกินกำหนด']
    ) if contract else Invoice.objects.none()
    unpaid_total    = unpaid_invoices.aggregate(total=django_models.Sum('Grand_Total'))['total'] or Decimal('0')
    unpaid_count    = unpaid_invoices.count()

    deposit         = contract.Deposit          if contract else Decimal('0')
    deposit_advance = contract.Deposit_Advance  if contract else Decimal('0')
    total_collateral = deposit + deposit_advance  # เงินประกัน + เงินมัดจำ

    # เคส 1: ไม่มียอดค้าง
    # เคส 2: ค้าง ≤ ประกัน+มัดจำ → หักได้เลย
    # เคส 3: ค้าง > ประกัน+มัดจำ → ต้องโปะเพิ่ม + บันทึกเหตุผล
    can_deduct  = unpaid_total > 0 and unpaid_total <= total_collateral
    need_extra  = unpaid_total > total_collateral
    refund      = total_collateral - unpaid_total if can_deduct else Decimal('0')
    shortfall   = unpaid_total - total_collateral if need_extra else Decimal('0')

    ctx_base = {
        'room': room, 'contract': contract, 'action': 'moveout',
        'unpaid_count': unpaid_count, 'unpaid_total': unpaid_total,
        'deposit': deposit, 'deposit_advance': deposit_advance,
        'total_collateral': total_collateral,
        'can_deduct': can_deduct, 'refund': refund,
        'need_extra': need_extra, 'shortfall': shortfall,
    }

    if request.method == 'POST':
        note        = request.POST.get('moveout_note', '').strip()
        extra_str   = request.POST.get('extra_payment', '0').strip() or '0'

        try:
            extra_payment = Decimal(extra_str)
        except Exception:
            extra_payment = Decimal('0')

        # Validation: ต้องมีเหตุผลเสมอถ้าค้างเกินประกัน+มัดจำ
        if need_extra and not note:
            return render(request, 'apartment/room/action_confirm.html',
                          {**ctx_base, 'error': 'กรุณาระบุเหตุผล / ข้อตกลงก่อนดำเนินการ'})

        # Validation: ยอดโปะ + ประกัน+มัดจำ ต้องครบยอดค้าง
        if need_extra and (extra_payment + total_collateral) < unpaid_total:
            remaining = unpaid_total - total_collateral - extra_payment
            return render(request, 'apartment/room/action_confirm.html',
                          {**ctx_base, 'error': f'ยอดรวมยังไม่ครบ ขาดอีก {remaining:,.2f} บาท', 'extra_payment': extra_str})

        if contract:
            contract.Status = 'หมดอายุ'
            # บันทึก note รวมยอดโปะเพิ่ม (ถ้ามี)
            full_note = note
            if need_extra and extra_payment > 0:
                full_note += f'\n[ผู้เช่าจ่ายเพิ่ม {extra_payment:,.2f} บาท + หักประกัน/มัดจำ {total_collateral:,.2f} บาท]'
            if full_note:
                contract.Moveout_Note = full_note
            contract.save()

            # mark invoice ค้างทั้งหมดเป็น ชำระแล้ว (หักประกัน หรือโปะเพิ่มจนครบ)
            if unpaid_total > 0 and (can_deduct or (need_extra and (extra_payment + total_collateral) >= unpaid_total)):
                unpaid_invoices.update(Status='ชำระแล้ว', Paid_Date=datetime.date.today())

        room.Status      = 'ว่าง'
        room.Status_Flag = 'รอทำความสะอาด'
        room.save()
        return redirect('room_detail', pk=pk)

    return render(request, 'apartment/room/action_confirm.html', ctx_base)


@login_required
@role_required('ADMIN', 'MANAGER')
def room_action_notify_out(request, pk):
    # แจ้งย้ายออก: เปลี่ยน Status_Flag เป็น แจ้งย้ายออก
    room = get_object_or_404(Room, pk=pk)

    if request.method == 'POST':
        room.Status_Flag = 'แจ้งย้ายออก'
        room.save()
        return redirect('room_detail', pk=pk)

    return render(request, 'apartment/room/action_confirm.html', {
        'room':    room,
        'action':  'notify_out',
        'title':   f'แจ้งย้ายออก — ห้อง {room.Room_Number}',
        'message': f'บันทึกว่าผู้เช่าห้อง {room.Room_Number} แจ้งความประสงค์จะย้ายออก ?',
        'btn_color': 'warning',
    })


@login_required
@role_required('ADMIN', 'MANAGER')
def room_action_clean(request, pk):
    # แจ้งทำความสะอาด
    room = get_object_or_404(Room, pk=pk)

    if request.method == 'POST':
        room.Status_Flag = 'รอทำความสะอาด'
        room.save()
        return redirect('room_detail', pk=pk)

    return render(request, 'apartment/room/action_confirm.html', {
        'room':    room,
        'action':  'clean',
        'title':   f'แจ้งทำความสะอาด — ห้อง {room.Room_Number}',
        'message': f'บันทึกว่าห้อง {room.Room_Number} ต้องการทำความสะอาด ?',
        'btn_color': 'info',
    })


@login_required
@role_required('ADMIN', 'MANAGER')
def room_action_done_clean(request, pk):
    # ทำความสะอาดเสร็จ → คืนสถานะปกติ
    room = get_object_or_404(Room, pk=pk)
    if request.method == 'POST':
        room.Status_Flag = 'ปกติ'
        room.save()
        return redirect('room_detail', pk=pk)
    return render(request, 'apartment/room/action_confirm.html', {
        'room':    room,
        'action':  'done_clean',
        'title':   f'ทำความสะอาดเสร็จ — ห้อง {room.Room_Number}',
        'message': f'ยืนยันว่าห้อง {room.Room_Number} ทำความสะอาดเสร็จแล้ว ?',
        'btn_color': 'success',
    })


@login_required
@role_required('ADMIN', 'MANAGER')
def invoice_send_email(request, pk):
    invoice      = get_object_or_404(Invoice, pk=pk)
    monthly_bill = MonthlyBill.objects.filter(Invoice_ID=invoice).first()
    utility      = Utility.objects.filter(Invoice_ID=invoice).first()
    fines        = Fine.objects.filter(Invoice_ID=invoice)
    tenant       = invoice.Contract_ID.Tenant_ID

    # ตรวจสอบว่ามีอีเมลไหม
    if not tenant.Email:
        return render(request, 'apartment/invoice/email_result.html', {
            'success': False,
            'message': f'ผู้เช่า {tenant.First_Name} {tenant.Last_Name} ไม่มีอีเมลในระบบ',
            'invoice': invoice,
        })

    if request.method == 'POST':
        # render HTML สำหรับส่งเป็น email body
        email_body = render_to_string('apartment/invoice/email_body.html', {
            'invoice':      invoice,
            'monthly_bill': monthly_bill,
            'utility':      utility,
            'fines':        fines,
            'tenant':       tenant,
        })

        try:
            send_mail(
                subject  = f'ใบแจ้งหนี้ห้อง {invoice.Contract_ID.Room_ID} — เดือน {invoice.Billing_Date.strftime("%B %Y")}',
                message  = '',                  # plain text (เว้นว่างเพราะใช้ html)
                from_email = None,              # ใช้ DEFAULT_FROM_EMAIL
                recipient_list = [tenant.Email],
                html_message = email_body,
                fail_silently = False,
            )
            return render(request, 'apartment/invoice/email_result.html', {
                'success': True,
                'message': f'ส่งอีเมลไปที่ {tenant.Email} เรียบร้อยแล้ว',
                'invoice': invoice,
            })
        except Exception as e:
            return render(request, 'apartment/invoice/email_result.html', {
                'success': False,
                'message': f'ส่งอีเมลไม่สำเร็จ: {str(e)}',
                'invoice': invoice,
            })

    # GET: หน้ายืนยันก่อนส่ง
    return render(request, 'apartment/invoice/email_confirm.html', {
        'invoice': invoice,
        'tenant':  tenant,
    })
